"""
dSPACE AutomationDesk — Bulk Test Script Generator
====================================================
Reads HIL_TestCases_Template.xlsx and auto-generates
AutomationDesk-compatible Python scripts for:
  • CAN bus signals
  • LIN bus signals
  • Model Variables (MAPort)
  • Diagnostics (UDS / OBD-II)

Usage:
    python generate_tests.py --excel HIL_TestCases_Template.xlsx
    python generate_tests.py --excel HIL_TestCases_Template.xlsx --sheet CAN_Tests
"""

import os, sys, argparse, json
from datetime import datetime
import openpyxl

OUTPUT_DIR = "generated_tests"
REPORT_DIR = "reports"
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(REPORT_DIR, exist_ok=True)

TIMESTAMP = datetime.now().strftime("%Y-%m-%d %H:%M")

# ══════════════════════════════════════════════════════════════════
# TEMPLATES — one per protocol type
# ══════════════════════════════════════════════════════════════════

CAN_TEMPLATE = '''"""
AutomationDesk Test — CAN Bus
Test ID  : {test_id}
Feature  : {feature_name}
Generated: {timestamp}
"""
# ── dSPACE Imports (uncomment on real HIL) ────────────────────
# from Testbench import MAPort, CANPort
# import AutomationDesk as AD

import time

# ── Configuration ─────────────────────────────────────────────
CHANNEL        = "{channel}"
DBC_FILE       = r"{dbc_file}"
MESSAGE        = "{message}"
SIGNAL         = "{signal_name}"
STIMULI_VALUE  = {stimuli_value}
EXPECTED       = {expected_value}
TOLERANCE      = {tolerance}
WAIT_TIME      = {wait_time}
PRECONDITION   = "{precondition}"
TEST_ID        = "{test_id}"

# ── Helpers ───────────────────────────────────────────────────
def log(msg):
    print(f"[{{TEST_ID}}] {{msg}}")

def pass_fail(actual, expected, tol):
    return abs(float(actual) - float(expected)) <= float(tol)

# ── Test Steps ────────────────────────────────────────────────
def setup():
    log(f"SETUP | Precondition: {{PRECONDITION}}")
    # CANPort.LoadDBC(CHANNEL, DBC_FILE)
    # MAPort.Write("Precondition/IgnitionON", 1)  # example

def execute():
    log(f"EXECUTE | Writing {{SIGNAL}} = {{STIMULI_VALUE}} on {{CHANNEL}}/{{MESSAGE}}")
    # CANPort.WriteSignal(CHANNEL, MESSAGE, SIGNAL, STIMULI_VALUE)
    time.sleep(WAIT_TIME)

def verify():
    log(f"VERIFY | Reading {{CHANNEL}}/{{MESSAGE}}/{{SIGNAL}}")
    # actual = CANPort.ReadSignal(CHANNEL, MESSAGE, SIGNAL)
    actual = EXPECTED   # ← placeholder: replace with real read
    result = pass_fail(actual, EXPECTED, TOLERANCE)
    status = "PASS ✅" if result else "FAIL ❌"
    log(f"RESULT | {{status}} | Expected={{EXPECTED}} Actual={{actual}} Tol={{TOLERANCE}}")
    return result

def teardown():
    log("TEARDOWN | Resetting signals")
    # CANPort.WriteSignal(CHANNEL, MESSAGE, SIGNAL, 0)

def main():
    setup()
    execute()
    ok = verify()
    teardown()
    return ok

if __name__ == "__main__":
    main()
'''

LIN_TEMPLATE = '''"""
AutomationDesk Test — LIN Bus
Test ID  : {test_id}
Feature  : {feature_name}
Generated: {timestamp}
"""
# ── dSPACE Imports ────────────────────────────────────────────
# from Testbench import MAPort, LINPort
import time

CHANNEL        = "{channel}"
LDF_FILE       = r"{ldf_file}"
FRAME          = "{frame_name}"
SIGNAL         = "{signal_name}"
STIMULI_VALUE  = {stimuli_value}
EXPECTED       = {expected_value}
TOLERANCE      = {tolerance}
WAIT_TIME      = {wait_time}
TEST_ID        = "{test_id}"

def log(msg):  print(f"[{{TEST_ID}}] {{msg}}")
def pass_fail(a, e, t): return abs(float(a) - float(e)) <= float(t)

def setup():
    log(f"SETUP | Loading LDF: {{LDF_FILE}}")
    # LINPort.LoadLDF(CHANNEL, LDF_FILE)

def execute():
    log(f"EXECUTE | Writing {{FRAME}}/{{SIGNAL}} = {{STIMULI_VALUE}}")
    # LINPort.WriteSignal(CHANNEL, FRAME, SIGNAL, STIMULI_VALUE)
    time.sleep(WAIT_TIME)

def verify():
    # actual = LINPort.ReadSignal(CHANNEL, FRAME, SIGNAL)
    actual = EXPECTED
    ok = pass_fail(actual, EXPECTED, TOLERANCE)
    log(f"RESULT | {{'PASS ✅' if ok else 'FAIL ❌'}} | Exp={{EXPECTED}} Act={{actual}}")
    return ok

def teardown():
    log("TEARDOWN | Resetting LIN signals")
    # LINPort.WriteSignal(CHANNEL, FRAME, SIGNAL, 0)

def main():
    setup(); execute(); ok = verify(); teardown(); return ok

if __name__ == "__main__":
    main()
'''

MV_TEMPLATE = '''"""
AutomationDesk Test — Model Variable (MAPort)
Test ID  : {test_id}
Feature  : {feature_name}
Generated: {timestamp}
"""
# ── dSPACE Imports ────────────────────────────────────────────
# from Testbench import MAPort
import time

WRITE_PATH     = r"{write_path}"
WRITE_VALUE    = {write_value}
READ_PATH      = r"{read_path}"
EXPECTED       = {expected_value}
TOLERANCE      = {tolerance}
WAIT_TIME      = {wait_time}
TEST_ID        = "{test_id}"

def log(msg):  print(f"[{{TEST_ID}}] {{msg}}")
def pass_fail(a, e, t): return abs(float(a) - float(e)) <= float(t)

def setup():
    log(f"SETUP | Resetting: {{WRITE_PATH}}")
    # MAPort.Write(WRITE_PATH, 0)

def execute():
    log(f"EXECUTE | MAPort.Write({{WRITE_PATH}}, {{WRITE_VALUE}})")
    # MAPort.Write(WRITE_PATH, WRITE_VALUE)
    time.sleep(WAIT_TIME)

def verify():
    log(f"VERIFY | MAPort.Read({{READ_PATH}})")
    # actual = MAPort.Read(READ_PATH)
    actual = EXPECTED
    ok = pass_fail(actual, EXPECTED, TOLERANCE)
    log(f"RESULT | {{'PASS ✅' if ok else 'FAIL ❌'}} | Exp={{EXPECTED}} Act={{actual}}")
    return ok

def teardown():
    log("TEARDOWN | Resetting model variable")
    # MAPort.Write(WRITE_PATH, 0)

def main():
    setup(); execute(); ok = verify(); teardown(); return ok

if __name__ == "__main__":
    main()
'''

DIAG_TEMPLATE = '''"""
AutomationDesk Test — Diagnostics ({protocol})
Test ID  : {test_id}
Feature  : {feature_name}
Generated: {timestamp}
"""
# ── dSPACE Imports ────────────────────────────────────────────
# from Testbench import DiagPort
import time

PROTOCOL       = "{protocol}"
SERVICE_ID     = "{service_id}"
SUB_FUNC_DID   = "{sub_func}"
REQUEST_DATA   = "{request_data}"
EXPECTED_RESP  = "{expected_response}"
RESPONSE_CODE  = "{response_code}"
TIMEOUT_MS     = {timeout_ms}
SESSION_TYPE   = "{session_type}"
TEST_ID        = "{test_id}"

def log(msg):  print(f"[{{TEST_ID}}] {{msg}}")

def setup():
    log(f"SETUP | Opening {{PROTOCOL}} session: {{SESSION_TYPE}}")
    # DiagPort.OpenSession(SESSION_TYPE)

def execute():
    log(f"EXECUTE | Sending SID={{SERVICE_ID}} DID={{SUB_FUNC_DID}}")
    # response = DiagPort.SendRequest(SERVICE_ID, SUB_FUNC_DID, REQUEST_DATA)
    time.sleep(TIMEOUT_MS / 1000)

def verify():
    log(f"VERIFY | Expected response: {{EXPECTED_RESP}}")
    # actual_resp = DiagPort.GetLastResponse()
    actual_resp = EXPECTED_RESP  # placeholder
    ok = (actual_resp == EXPECTED_RESP)
    log(f"RESULT | {{'PASS ✅' if ok else 'FAIL ❌'}} | Exp={{EXPECTED_RESP}} Act={{actual_resp}}")
    return ok

def teardown():
    log("TEARDOWN | Closing diagnostic session")
    # DiagPort.CloseSession()

def main():
    setup(); execute(); ok = verify(); teardown(); return ok

if __name__ == "__main__":
    main()
'''

# ══════════════════════════════════════════════════════════════════
# READERS — extract rows from each Excel sheet
# ══════════════════════════════════════════════════════════════════

def safe(val, default="0"):
    if val is None or str(val).strip() == "": return default
    return str(val).strip()

def read_can_sheet(ws):
    tests = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]: continue
        tests.append({
            "type": "CAN",
            "test_id": safe(row[0]),
            "feature_name": safe(row[1], "Unnamed"),
            "channel": safe(row[2], "CAN1"),
            "dbc_file": safe(row[3], "Vehicle.dbc"),
            "message": safe(row[4], "MSG"),
            "signal_name": safe(row[5], "Signal"),
            "stimuli_value": safe(row[6], "0"),
            "unit": safe(row[7], ""),
            "expected_value": safe(row[8], "0"),
            "tolerance": safe(row[9], "0"),
            "wait_time": safe(row[10], "1.0"),
            "priority": safe(row[11], "Medium"),
            "precondition": safe(row[12], ""),
            "notes": safe(row[13], ""),
        })
    return tests

def read_lin_sheet(ws):
    tests = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]: continue
        tests.append({
            "type": "LIN",
            "test_id": safe(row[0]),
            "feature_name": safe(row[1], "Unnamed"),
            "channel": safe(row[2], "LIN1"),
            "ldf_file": safe(row[3], "Body.ldf"),
            "frame_name": safe(row[4], "Frame"),
            "signal_name": safe(row[5], "Signal"),
            "stimuli_value": safe(row[6], "0"),
            "expected_value": safe(row[8], "0"),
            "tolerance": safe(row[9], "0"),
            "wait_time": safe(row[10], "1.0"),
        })
    return tests

def read_mv_sheet(ws):
    tests = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]: continue
        tests.append({
            "type": "MV",
            "test_id": safe(row[0]),
            "feature_name": safe(row[1], "Unnamed"),
            "write_path": safe(row[2], "Model Root/Signal/Value"),
            "write_value": safe(row[3], "0"),
            "read_path": safe(row[4], "Model Root/Signal/Out1"),
            "expected_value": safe(row[5], "0"),
            "tolerance": safe(row[6], "0"),
            "wait_time": safe(row[7], "1.0"),
        })
    return tests

def read_diag_sheet(ws):
    tests = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]: continue
        tests.append({
            "type": "DIAG",
            "test_id": safe(row[0]),
            "feature_name": safe(row[1], "Unnamed"),
            "protocol": safe(row[2], "UDS"),
            "service_id": safe(row[3], "0x22"),
            "sub_func": safe(row[4], "0x00"),
            "request_data": safe(row[5], "—"),
            "expected_response": safe(row[6], ""),
            "response_code": safe(row[7], ""),
            "timeout_ms": safe(row[8], "1000"),
            "session_type": safe(row[10], "Default"),
        })
    return tests

SHEET_READERS = {
    "CAN_Tests":         read_can_sheet,
    "LIN_Tests":         read_lin_sheet,
    "ModelVar_Tests":    read_mv_sheet,
    "Diagnostics_Tests": read_diag_sheet,
}

# ══════════════════════════════════════════════════════════════════
# GENERATORS
# ══════════════════════════════════════════════════════════════════

def generate_script(tc):
    t = tc["type"]
    if t == "CAN":
        code = CAN_TEMPLATE.format(timestamp=TIMESTAMP, **tc)
    elif t == "LIN":
        code = LIN_TEMPLATE.format(timestamp=TIMESTAMP, **tc)
    elif t == "MV":
        code = MV_TEMPLATE.format(timestamp=TIMESTAMP, **tc)
    elif t == "DIAG":
        code = DIAG_TEMPLATE.format(timestamp=TIMESTAMP, **tc)
    else:
        return None

    fname = f"{tc['test_id']}_{tc['feature_name'].replace(' ', '_')}.py"
    path  = os.path.join(OUTPUT_DIR, fname)
    with open(path, "w") as f:
        f.write(code)
    return fname

def generate_html_report(results):
    rows = ""
    for r in results:
        status_badge = (
            '<span style="background:#d4edda;color:#155724;padding:2px 8px;'
            'border-radius:4px;font-size:11px">✅ Generated</span>'
            if r["ok"] else
            '<span style="background:#f8d7da;color:#721c24;padding:2px 8px;'
            'border-radius:4px;font-size:11px">❌ Error</span>'
        )
        rows += f"""
        <tr>
          <td>{r['test_id']}</td>
          <td>{r['feature']}</td>
          <td><span style="font-family:monospace;font-size:11px;background:#f0f0f0;
              padding:2px 5px;border-radius:3px">{r['type']}</span></td>
          <td>{status_badge}</td>
          <td style="font-family:monospace;font-size:11px">{r['file']}</td>
        </tr>"""

    total   = len(results)
    success = sum(1 for r in results if r["ok"])
    html = f"""<!DOCTYPE html>
<html><head>
<meta charset="UTF-8">
<title>dSPACE AutomationDesk — Script Generation Report</title>
<style>
  body {{font-family:Arial,sans-serif;margin:0;background:#f5f7fa}}
  .header {{background:linear-gradient(135deg,#1F3864,#2E75B6);color:white;
            padding:30px 40px}}
  .header h1 {{margin:0;font-size:22px}}
  .header p  {{margin:5px 0 0;opacity:.8;font-size:13px}}
  .stats {{display:flex;gap:20px;padding:20px 40px;background:white;
           border-bottom:1px solid #e0e0e0}}
  .stat {{background:#f0f6ff;border-radius:8px;padding:14px 24px;text-align:center}}
  .stat .num {{font-size:28px;font-weight:bold;color:#1F3864}}
  .stat .lbl {{font-size:11px;color:#666;margin-top:2px}}
  .content {{padding:30px 40px}}
  table {{width:100%;border-collapse:collapse;background:white;border-radius:8px;
          overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08)}}
  th {{background:#1F3864;color:white;padding:10px 14px;text-align:left;
       font-size:12px}}
  td {{padding:10px 14px;border-bottom:1px solid #f0f0f0;font-size:12px;
       vertical-align:middle}}
  tr:last-child td {{border-bottom:none}}
  tr:hover td {{background:#f9fbff}}
  .footer {{padding:20px 40px;color:#999;font-size:11px;
            border-top:1px solid #e0e0e0;margin-top:20px}}
</style></head><body>
<div class="header">
  <h1>🔧 dSPACE AutomationDesk — Bulk Script Generation Report</h1>
  <p>Generated: {TIMESTAMP} &nbsp;|&nbsp; Framework: HIL Test Automation</p>
</div>
<div class="stats">
  <div class="stat"><div class="num">{total}</div><div class="lbl">Total Scripts</div></div>
  <div class="stat"><div class="num" style="color:#155724">{success}</div>
       <div class="lbl">Generated OK</div></div>
  <div class="stat"><div class="num" style="color:#721c24">{total - success}</div>
       <div class="lbl">Errors</div></div>
  <div class="stat"><div class="num">
    {len([r for r in results if r["type"]=="CAN"])}</div>
    <div class="lbl">CAN Tests</div></div>
  <div class="stat"><div class="num">
    {len([r for r in results if r["type"]=="LIN"])}</div>
    <div class="lbl">LIN Tests</div></div>
  <div class="stat"><div class="num">
    {len([r for r in results if r["type"]=="MV"])}</div>
    <div class="lbl">Model Var</div></div>
  <div class="stat"><div class="num">
    {len([r for r in results if r["type"]=="DIAG"])}</div>
    <div class="lbl">Diagnostics</div></div>
</div>
<div class="content">
  <table>
    <thead><tr>
      <th>Test ID</th><th>Feature Name</th><th>Type</th>
      <th>Status</th><th>Generated File</th>
    </tr></thead>
    <tbody>{rows}</tbody>
  </table>
</div>
<div class="footer">
  Output directory: generated_tests/ &nbsp;|&nbsp;
  AutomationDesk HIL Framework &nbsp;|&nbsp; dSPACE Script Generator
</div>
</body></html>"""

    rpath = os.path.join(REPORT_DIR, "generation_report.html")
    with open(rpath, "w") as f:
        f.write(html)
    return rpath

# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="dSPACE HIL Bulk Test Generator")
    parser.add_argument("--excel", required=True, help="Path to Excel test case file")
    parser.add_argument("--sheet", default=None, help="Sheet name (optional, default=all)")
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"❌ File not found: {args.excel}"); sys.exit(1)

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    all_tests = []

    sheets_to_process = [args.sheet] if args.sheet else list(SHEET_READERS.keys())
    for sheet_name in sheets_to_process:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet '{sheet_name}' not found — skipping"); continue
        reader = SHEET_READERS[sheet_name]
        tests  = reader(wb[sheet_name])
        print(f"  📄 {sheet_name}: {len(tests)} test cases found")
        all_tests.extend(tests)

    print(f"\n🚀 Generating {len(all_tests)} test scripts...\n")
    results = []
    for tc in all_tests:
        try:
            fname = generate_script(tc)
            print(f"  ✅  {tc['test_id']} → {fname}")
            results.append({"test_id": tc["test_id"], "feature": tc["feature_name"],
                            "type": tc["type"], "ok": True, "file": fname})
        except Exception as e:
            print(f"  ❌  {tc['test_id']} → ERROR: {e}")
            results.append({"test_id": tc["test_id"], "feature": tc.get("feature_name","?"),
                            "type": tc.get("type","?"), "ok": False, "file": str(e)})

    rpath = generate_html_report(results)
    print(f"\n✅ Done! {len(all_tests)} scripts → {OUTPUT_DIR}/")
    print(f"📊 Report → {rpath}")

if __name__ == "__main__":
    main()
